import openpyxl as opx

workbook = opx.load_workbook('Prepared_Data.xlsx')
#for chain dump the sequence into a file called rownum.fasta
sh_train_chains = workbook['Train_Chains'];
for i in range(2,sh_train_chains.max_row + 1):
    pdbid = sh_train_chains.cell(row=i,column=1).value
    chain_id = str(sh_train_chains.cell(row=i,column=2).value)
    seq = sh_train_chains.cell(row=i,column=8).value
    filename='fastas\\'+str(i)+'.fasta'
    line1 = '>' + pdbid + '-' + chain_id
    line2 = seq
    f = open(filename,'w')
#    f.write(line1+'\n')
    f.write(line2+'\n')
    f.close()
    print('pdbid:'+pdbid+' chainid:'+chain_id+' file name:'+filename+'\n')
