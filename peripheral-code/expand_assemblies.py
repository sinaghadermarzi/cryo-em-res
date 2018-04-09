

import openpyxl as opx

workbook = opx.load_workbook('Prepared_Data.xlsx')
#sh_assemblies = workbook['Assemblies(2015-2017)']
sh_chains = workbook['Chains(2015-2017)']
sh_train_chains = workbook['Train_Chains']
sh_train_set = workbook['Train_Set']
output_current_row = 2 
chain_row = 1
for i in range(2,sh_train_set.max_row + 1):

    query_id = sh_train_set.cell(row=i,column=1).value


    found = 0
    print('looking for entry '+str(i)+': '+query_id+'\n')
    while (chain_row <= sh_chains.max_row) and (found==0):#look at the chain table row by row until you reach an entry with the same id as query id
        curr_id = sh_chains.cell(row=chain_row,column=1).value
        if curr_id == query_id:
            found = 1
            print('found')
        else:
            chain_row = chain_row+1
    if found == 1:
        while (chain_row <= sh_chains.max_row) and (curr_id == query_id):
            for k in range(1,sh_chains.max_column+1):
                sh_train_chains.cell(row=output_current_row,column=k).value = sh_chains.cell(row=chain_row,column=k).value
                print(k)

            output_current_row = output_current_row+1
            chain_row = chain_row + 1
            curr_id = sh_chains.cell(row=chain_row,column=1).value

    else:
        print('entry '+str(i)+'not found')

workbook.template = False
workbook.save('Prepared_Data.xlsx')